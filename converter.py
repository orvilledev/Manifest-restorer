"""Convert Amazon shipment manifest CSV to Send-to-Amazon Excel template."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from pathlib import Path

import openpyxl


@dataclass
class ManifestMetadata:
    shipment_number: str = ""
    workflow_name: str = ""
    sku_count: int = 0
    unit_count: int = 0


@dataclass
class ManifestItem:
    merchant_sku: str
    quantity: int


def _strip_quotes(value: str) -> str:
    return value.strip().strip('"')


def parse_manifest_csv(content: str | bytes) -> tuple[ManifestMetadata, list[ManifestItem]]:
    """Parse an Amazon shipment manifest CSV export."""
    if isinstance(content, bytes):
        text = content.decode("utf-8-sig")
    else:
        text = content

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    metadata = ManifestMetadata()
    items: list[ManifestItem] = []
    in_data_section = False

    for row in rows:
        if not row:
            continue

        first = _strip_quotes(row[0])

        if not in_data_section:
            if first == "Shipment number" and len(row) > 1:
                metadata.shipment_number = _strip_quotes(row[1])
            elif first == "Workflow name" and len(row) > 1:
                metadata.workflow_name = _strip_quotes(row[1])
            elif first == "SKUs" and len(row) > 1:
                metadata.sku_count = int(_strip_quotes(row[1]) or 0)
            elif first == "Units" and len(row) > 1:
                metadata.unit_count = int(_strip_quotes(row[1]) or 0)
            elif first == "SKU":
                in_data_section = True
            continue

        if "Box ID" in str(row):
            break

        if len(row) < 8 or not first:
            continue

        quantity_raw = _strip_quotes(row[7])
        items.append(
            ManifestItem(
                merchant_sku=first,
                quantity=int(quantity_raw) if quantity_raw else 0,
            )
        )

    return metadata, items


def _find_template_sheet(workbook: openpyxl.Workbook) -> str:
    for name in workbook.sheetnames:
        if "template" in name.lower():
            return name
    raise ValueError("Template sheet not found in workbook.")


def convert_to_excel(
    csv_content: str | bytes,
    template_path: str | Path,
) -> tuple[bytes, ManifestMetadata, list[ManifestItem]]:
    """Convert manifest CSV content into a populated Excel template."""
    metadata, items = parse_manifest_csv(csv_content)

    workbook = openpyxl.load_workbook(template_path)
    sheet_name = _find_template_sheet(workbook)
    worksheet = workbook[sheet_name]

    # Clear existing SKU rows (data starts at row 7).
    if worksheet.max_row >= 7:
        worksheet.delete_rows(7, worksheet.max_row - 6)

    for item in items:
        worksheet.append(
            [
                item.merchant_sku,
                item.quantity,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue(), metadata, items
